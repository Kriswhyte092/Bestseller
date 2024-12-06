from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import pandas as pd
import csv
import re


class GoalExtractor:
    def __init__(self, file_path, output_csv):
        self.file_path = file_path
        self.output_csv = output_csv
        self.name_mapping = {
            "VMK": "Vero Moda Kringlan",
            "VIK": "Vila Kringlan",
            "NIK": "Name It Kringlan",
            "JJK": "Jack & Jones Kringlan",
            "SLK": "Selected Kringlan",
            "VMS": "Vero Moda Smáralind",
            "VIS": "Vila Smáralind",
            "NIS": "Name It Smáralind",
            "JJS": "Jack & Jones Smáralind",
            "SLS": "Selected Smáralind",
        }

    def format_date(self, value):
        if isinstance(value, datetime):
            return value.strftime("%d %b").lstrip("0")
        elif isinstance(value, str):
            try:
                date_obj = datetime.strptime(value, "%d.%m.%Y")
                return date_obj.strftime("%d %b").lstrip("0")
            except ValueError:
                return value
        return value

    def read_excel(self, start_row, end_row):
        wb = load_workbook(self.file_path)
        ws = wb.active

        results = {}
        for row_number in range(start_row, end_row + 1):
            row_values = [cell.value for cell in ws[row_number]]
            name = self.name_mapping.get(row_values[0], row_values[0])
            grouped_values = []
            for i in range(1, len(row_values), 3):
                if i + 1 < len(row_values):
                    try:
                        val1 = int(row_values[i])
                        val2 = int(row_values[i + 1])
                        if val1 != 0:
                            ratio = round(val2 / val1, 2)
                        else:
                            ratio = None
                        grouped_values.append([ratio])
                    except (ValueError, TypeError):
                        grouped_values.append([None])
            results[name] = grouped_values

        wb.close()
        return results

    def print_second_line(self):
        wb = load_workbook(self.file_path)
        ws = wb.active

        row_2_values = [cell.value for cell in ws[2]]
        row_2_values = row_2_values[1:]
        filtered_row_2_values = [value for value in row_2_values if value is not None]
        formatted_values = [self.format_date(value) for value in filtered_row_2_values]

        wb.close()
        return formatted_values

    def get_bonus(self, name, index):
        thresholds = [
            (0.8, 2000),
            (1.0, 2000),
            (1.15, 3000),
            (1.3, 4000),
            (1.5, 5000),
            (1.75, 7500),
            (2.0, 10000),
        ]

        special_stores = {"VMK", "VIK", "NIK", "SLK"}
        bonus = 0
        if name in special_stores:
            for threshold, value in thresholds:
                if index >= threshold:
                    bonus = value
        else:
            for threshold, value in thresholds[1:]:
                if index >= threshold:
                    bonus = value
        return bonus

    def pair_and_print(self):
        second_line_values = self.print_second_line()
        excel_results = self.read_excel(4, 13)

        formatted_result = []
        for name, grouped_values in excel_results.items():
            paired_values = list(zip(second_line_values, grouped_values))
            for date, index in paired_values:
                if index[0] is None:
                    index[0] = 0
                bonus = self.get_bonus(name, index[0])
                if bonus != 0:
                    formatted_result.append(
                        {"date": self.format_date(date), "name": name, "index": bonus}
                    )

        with open(self.output_csv, mode="w", newline="", encoding="latin-1") as file:
            writer = csv.writer(file)
            for row in formatted_result:
                writer.writerow([row["date"], row["name"], row["index"]])

        print(f"The formatted result has been written to {self.output_csv}")


class EmployeeBonusProcessor:
    def __init__(self, shifts_file, bonus_file):
        """
        Initialize the processor with the shifts and bonuses file paths.
        """
        self.shifts_file = shifts_file
        self.bonus_file = bonus_file
        self.shifts_data = self._load_shifts_data()

    def _load_shifts_data(self):
        """
        Load the shifts data into a dictionary keyed by date.
        """
        shifts = pd.read_csv(self.shifts_file, encoding="utf-8")
        shifts_data = defaultdict(list)
        for day in shifts.columns:
            for row in shifts[day].dropna():
                date_match = re.match(r"(\d+\s\w+)", row)
                if date_match:
                    date = date_match.group(1)
                    shifts_data[date].append(row)
        return shifts_data

    def _preprocess_bonuses(self, bonus_file):
        """
        Preprocess the bonus file to ensure correct headers and formatting.
        """
        bonuses = pd.read_csv(bonus_file, encoding="latin1", header=None)

        # Check and rename columns dynamically if required
        if bonuses.shape[1] == 3:
            bonuses.columns = ["Date", "Store", "Bonus"]
        else:
            raise ValueError("Unexpected file structure in bonuses file.")

        # Clean and normalize data
        bonuses["Date"] = bonuses["Date"].str.strip()
        bonuses["Store"] = bonuses["Store"].str.strip()
        bonuses["Bonus"] = bonuses["Bonus"].astype(float)

        return bonuses

    def _extract_employees(self, shift, store):
        """
        Extract employees working at a specific store in a shift entry.
        """
        employees = []
        lines = shift.split("\n")
        for i, line in enumerate(lines):
            if store in line:
                # The employee name is typically the following line after the store information
                if i + 1 < len(lines):
                    employee_name = lines[i + 1].strip()
                    if employee_name:
                        employees.append(employee_name)
        return employees

    def process_bonuses(self):
        """
        Process the bonuses from the bonuses file and assign them to employees.
        """
        bonuses = self._preprocess_bonuses(self.bonus_file)
        employee_bonuses = defaultdict(float)

        for _, row in bonuses.iterrows():
            date, store, bonus = row["Date"], row["Store"], row["Bonus"]
            if date in self.shifts_data:
                for shift in self.shifts_data[date]:
                    employees = self._extract_employees(shift, store)
                    for employee in employees:
                        employee_bonuses[employee] += bonus

        # Write the employee bonuses to a CSV file
        self.write_to_csv(employee_bonuses, "final_employee_bonuses.csv")
        # return employee_bonuses

    def write_to_csv(self, employee_bonuses, output_file):
        """
        Write the employee bonuses to a CSV file.
        """
        with open(output_file, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(["Employee", "Bonus"])
            for employee, bonus in employee_bonuses.items():
                writer.writerow([employee, bonus])
        print(f"The employee bonuses have been written to {output_file}")


# Stores indexes
file_path = input("excel: ")
# Output file for what days the stores have reached the indexes and the bonus amount
output_file = "bonus_dates_sorted.csv"
GoalExtractor(file_path, output_file).pair_and_print()

# Sling shift file path
shifts_file_path = input("sling csv: ")
bonus_file_path = "bonus_dates_sorted.csv"

EmployeeBonusProcessor(shifts_file_path, bonus_file_path).process_bonuses()

if __name__ == "__main__":
    # Stores indexes
    file_path = input("excel: ")
    output_file = "bonus_dates_sorted.csv"
    GoalExtractor(file_path, output_file).pair_and_print()

    # Sling shift file path
    shifts_file_path = input("sling csv: ")
    bonus_file_path = "bonus_dates_sorted.csv"

    EmployeeBonusProcessor(shifts_file_path, bonus_file_path).process_bonuses()
