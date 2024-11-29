from openpyxl import load_workbook
from datetime import datetime


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return value


def read_excel(file_path):
    # Load the workbook
    wb = load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active

    # Access the entire row 4
    row_number = 4
    row_values = [cell.value for cell in ws[row_number]]

    # Create a dictionary with the first value as the key and the rest grouped in threes
    name = row_values[0]
    grouped_values = []
    for i in range(1, len(row_values), 3):
        if i + 1 < len(
            row_values
        ):  # Ensure there are at least two values to form a ratio
            try:
                val1 = int(row_values[i])
                val2 = int(row_values[i + 1])
                if val1 != 0:  # Avoid division by zero
                    ratio = round(val2 / val1, 2)
                else:
                    ratio = None  # Handle division by zero case
                grouped_values.append([ratio])
            except (ValueError, TypeError):
                grouped_values.append([None])  # Handle non-integer values

    result = {name: grouped_values}

    # Close the workbook
    wb.close()

    return result


def print_second_line(file_path):
    # Load the workbook
    wb = load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active

    # Access the entire row 2
    row_2_values = [cell.value for cell in ws[2]]

    # Remove the first value from row 2
    row_2_values = row_2_values[1:]

    # Filter out rows where the corresponding value in row 2 is None
    filtered_row_2_values = [value for value in row_2_values if value is not None]

    # Format dates consistently
    formatted_values = [format_date(value) for value in filtered_row_2_values]

    # Close the workbook
    wb.close()

    return formatted_values


def pair_and_print(file_path):
    second_line_values = print_second_line(file_path)
    excel_result = read_excel(file_path)

    # Extract the grouped values from the dictionary
    name = list(excel_result.keys())[0]
    grouped_values = excel_result[name]

    # Pair values from the second line with the grouped values from the fourth line
    paired_values = list(zip(second_line_values, grouped_values))

    # Create a list of dictionaries with the desired format
    formatted_result = [
        {"date": date, "name": name, "index": index} for date, index in paired_values
    ]

    print(f"The formatted result is: {formatted_result}")


# Example usage
file_path = "office/bonus_prjct/Solubonusar_Bestseller_okt2024.xlsx"
pair_and_print(file_path)
