from openpyxl import load_workbook
from datetime import datetime
import csv

# Mapping of short names to full names
name_mapping = {
    "VMK": "Vero Moda Kringlan",
    "VIK": "Vila Kringlan",
    "NIK": "Name It Kringlan",
    "JJK": "Jack & Jones Kringlan",
    "SLK": "Selected Kringlan",
    "VMS": "Vero Moda Smáralind",
    "VIS": "Vila Smáralind",
    "NIS": "Name It Smáralind",
    "JJS": "Jack & Jones Smáralind",
    "SLS": "Selected Smáralind",
}


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d %b").lstrip("0")
    elif isinstance(value, str):
        try:
            date_obj = datetime.strptime(value, "%d.%m.%Y")
            return date_obj.strftime("%d %b").lstrip("0")
        except ValueError:
            return value
    return value


def read_excel(file_path, start_row, end_row):
    # Load the workbook
    wb = load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active

    results = {}
    for row_number in range(start_row, end_row + 1):
        row_values = [cell.value for cell in ws[row_number]]

        # Replace the short name with the full name
        name = name_mapping.get(row_values[0], row_values[0])

        grouped_values = []
        for i in range(1, len(row_values), 3):
            if i + 1 < len(
                row_values
            ):  # Ensure there are at least two values to form a ratio
                try:
                    val1 = int(row_values[i])
                    val2 = int(row_values[i + 1])
                    if val1 != 0:  # Avoid division by zero
                        ratio = round(val2 / val1, 2)
                    else:
                        ratio = None  # Handle division by zero case
                    grouped_values.append([ratio])
                except (ValueError, TypeError):
                    grouped_values.append([None])  # Handle non-integer values

        results[name] = grouped_values

    # Close the workbook
    wb.close()

    return results


def print_second_line(file_path):
    # Load the workbook
    wb = load_workbook(file_path)
    # Select the active worksheet
    ws = wb.active

    # Access the entire row 2
    row_2_values = [cell.value for cell in ws[2]]

    # Remove the first value from row 2
    row_2_values = row_2_values[1:]

    # Filter out rows where the corresponding value in row 2 is None
    filtered_row_2_values = [value for value in row_2_values if value is not None]

    # Format dates consistently
    formatted_values = [format_date(value) for value in filtered_row_2_values]

    # Close the workbook
    wb.close()

    return formatted_values


def get_bonus(name, index):
    # Define the thresholds and corresponding values
    thresholds = [
        (0.8, 2000),
        (1.0, 2000),
        (1.15, 3000),
        (1.3, 4000),
        (1.5, 5000),
        (1.75, 7500),
        (2.0, 10000),
    ]

    # Special stores
    special_stores = {"VMK", "VIK", "NIK", "SLK"}

    # Determine the bonus
    bonus = 0
    if name in special_stores:
        for threshold, value in thresholds:
            if index >= threshold:
                bonus = value
    else:
        for threshold, value in thresholds[1:]:  # Start from 1.0 for other stores
            if index >= threshold:
                bonus = value
    return bonus


def pair_and_print(file_path, output_csv):
    second_line_values = print_second_line(file_path)
    excel_results = read_excel(file_path, 4, 13)

    formatted_result = []
    for name, grouped_values in excel_results.items():
        # Pair values from the second line with the grouped values from the specified lines
        paired_values = list(zip(second_line_values, grouped_values))

        # Create a list of dictionaries with the desired format
        for date, index in paired_values:
            if index[0] is None:
                index[0] = 0  # Default to 0 if there is not a number in the index
            bonus = get_bonus(name, index[0])
            if bonus != 0:  # Only include entries with a non-zero bonus
                formatted_result.append(
                    {"date": format_date(date), "name": name, "index": bonus}
                )

    # Write the formatted result to a CSV file without the header
    with open(output_csv, mode="w", newline="") as file:
        writer = csv.writer(file)
        for row in formatted_result:
            writer.writerow([row["date"], row["name"], row["index"]])

    print(f"The formatted result has been written to {output_csv}")


# Example usage
file_path = "office/bonus_prjct/Solubonusar_Bestseller_okt2024.xlsx"
output_csv = "office/bonus_prjct/output.csv"
pair_and_print(file_path, output_csv)
