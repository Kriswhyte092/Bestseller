from openpyxl import load_workbook


def check_cell_value(item_number):
    existing_file = "retailItemList.xlsx"
    wb = load_workbook(existing_file)
    ws = wb.active

    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == item_number:
                # fær row number ef item_number matchar
                row_number = cell.row

                # Get item category code, or set default if empty
                itemCategoryCode = ws.cell(row=row_number, column=9).value
                if itemCategoryCode is None:
                    itemCategoryCode = "itemCategoryNotFound"

                # Get retail product code, or set default if empty
                retailProductCode = ws.cell(row=row_number, column=10).value
                if retailProductCode is None:
                    retailProductCode = "retailProductNotFound"

                wb.save(existing_file)
                wb.close()  # Close the workbook
                return itemCategoryCode, retailProductCode

    # If item number is not found
    wb.save(existing_file)
    wb.close()
    return "itemCategoryNotFound", "retailProductNotFound"
