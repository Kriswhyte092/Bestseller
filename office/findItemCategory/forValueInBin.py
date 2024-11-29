from openpyxl import load_workbook
import findItemCategoryCode


existing_file = #mögulega setja bara input hérna 
wb = load_workbook(existing_file)
ws = wb.active

#fer i gegnum hvert row i column 6=Fdálkur í excel og setur item number i fall
for row in ws.iter_rows(min_col=6, max_col=6):
    for cell in row:
        item_number = cell.value
        itemCategoryCode, retailProductCode = findItemCategoryCode.check_cell_value(
            item_number
        )
        row_number = cell.row
        ws.cell(row=row_number, column=10).value = itemCategoryCode
        ws.cell(row=row_number, column=11).value = retailProductCode

wb.save(existing_file)
