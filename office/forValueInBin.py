from openpyxl import load_workbook
import findItemCategoryCode


<<<<<<< HEAD
existing_file = 'E040.xlsx'
=======
existing_file = "A-hilla.xlsx"
>>>>>>> d310759c9e8a8db12fa43150a1ce56dfe976b0e8
wb = load_workbook(existing_file)
ws = wb.active


for row in ws.iter_rows(min_col=6, max_col=6):
    for cell in row:
        item_number = cell.value
        itemCategoryCode, retailProductCode = findItemCategoryCode.check_cell_value(
            item_number
        )
        row_number = cell.row
        ws.cell(row=row_number, column=10).value = itemCategoryCode
        ws.cell(row=row_number, column=11).value = retailProductCode


wb.save(existing_file)
