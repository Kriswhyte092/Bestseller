import openpyxl
import sqlite3
import shiftWrapper

class dataExtraction:
    def __init__(self, dbName, fileName):
        self.dbName = dbName
        self.fileName = fileName

    def openExcelFile(self):
        wb = openpyxl.load_workbook(self.fileName)
        ws = wb.active
        
        for row in ws.iter_rows(min_col=7, max_col=7):
            for cell in row:
                if cell.value != None:
                    rowNumber = cell.row
                    clockIn = cell.value
                    clockOut = ws.cell(row=rowNumber, column=12).value
                    name = ws.cell(row=rowNumber, column=2).value
                    id = ws.cell(row=rowNumber, column=4).value
                    date = ws.cell(row=rowNumber, column=3).value
                    location = ws.cell(row=rowNumber, column=6).value
                    status= ws.cell(row=rowNumber, column=17).value

                    self.ensureTableExists(name)
                    shift = shiftWrapper.shiftWrapper(name, id, date, location, clockIn, clockOut, status)
                    self.createShift(shift)


        wb.close()

    def createShift(self, shfit):
        tableName = self.mapName(shift.name)
        conn = sqlite3.connect(self.dbName)
        cursor = conn.cursor()

        cursor.execute(f"""
        INSERT INTO {tableName} (date, dv, ev)
        VALUES (?, ?, ?, ?, ?)
        """, (shift.date,
              shift.location, 
              shift.clockIn, 
              shift.clockOut, 
              shift.status))



    def mapName(self, name):
        return name.replace(" ", "_")
    
    def ensureTableExists(self, name):
        tableName = self.mapName(name) 
        conn = sqlite3.connect(self.dbName)
        cursor = conn.cursor()
    
        cursor.execute("""
        SELECT name FROM sqlite_master WHERE type='table' AND name=?;
        """, (tableName,))
        result = cursor.fetchone()
    
        if not result:
            cursor.execute(f"""
            CREATE TABLE {tableName} (
                shiftId INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                dv INTEGER
                ev INTEGER
            );
            """)
            print(f"Table '{tableName}' created.")
        else:
            print(f"Table '{tableName}' already exists.")
    
        conn.commit()
        conn.close()



data = dataExtraction("launaReiknivel.db", "test.xlsx")
data.openExcelFile()


