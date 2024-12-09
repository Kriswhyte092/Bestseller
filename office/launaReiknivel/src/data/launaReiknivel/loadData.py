import openpyxl
from modules.launaReiknivel.exportShift import exportShift
from datetime import datetime

class loadData:
    def __init__(self, fileName):
        self.fileName = fileName

    def openExcelFile(self):
        """Open the Excel file and process rows."""
        wb = openpyxl.load_workbook(self.fileName)
        ws = wb.active

            # Process each row starting from column 7 (clockIn data)
        for row in ws.iter_rows(min_row=2, max_col=17):  # Adjust row and column ranges as needed
            self.processRow(ws, row)
        wb.close()

    def processRow(self, ws, row):
        """Process a single row of the Excel sheet."""
        # Extract data from the row
        rowNumber = row[0].row
        name = ws.cell(row=rowNumber, column=2).value
        id = ws.cell(row=rowNumber, column=4).value
        date = ws.cell(row=rowNumber, column=3).value
        location = ws.cell(row=rowNumber, column=6).value
        clockIn = row[6].value  # Column 7 for clockIn
        clockOut = ws.cell(row=rowNumber, column=12).value
        status = ws.cell(row=rowNumber, column=17).value

        if self.cleanInput(date, clockIn, clockOut, status):
            shift = exportShift(name, id, date, location, clockIn, clockOut, status)
            shift.exportToDb()
        else:
            print("error in input data")

    def cleanInput(self, date, clockIn, clockOut, status):
        """Validate the input data."""
        return (
            self.cleanDate(date) and
            self.cleanClock(clockIn) and
            self.cleanClock(clockOut) and
            self.cleanStatus(status)
        )

    def cleanDate(self, date):
        if len(date) == 12:
            return True
        print("date not correct")
        return False
    
    def cleanClock(self, clock):
        try:
            datetime.strptime(clock, "%I:%M %p")
        except:
            print("clock not correct")
            return False
        return True

    def cleanStatus(self, status):
        if status == "Approved":
            return True
        print("status not approved")
        return False

